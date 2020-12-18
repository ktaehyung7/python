# -*- coding: utf-8 -*-
"""
Created on Thu Nov  5 14:22:40 2020

@author: Pulmuone
"""

import openpyxl
from openpyxl.styles import PatternFill, Color
import pandas as pd



wb = openpyxl.load_workbook("D:/Doc/12. IT 시스템 만족도 조사/IT만족도조사_exam/4_survey_all_2020(최종)v1.1.xlsx")

ws = wb.get_sheet_by_name("Sheet2")

print("max row = ", ws.max_row)
print("max col = ", ws.max_column)

sys_list = ['ebills', 'ERP', 'FMS', 'FNC', 'KWP', 'SFA', '톡톡']

for sys in sys_list:
    f_name = "D:/Doc/12. IT 시스템 만족도 조사/IT만족도조사_exam/R_data/모바일의견/원본/모바일의견("+sys+").txt"
    with open(f_name, 'w') as file1:
        for row1 in ws.rows:
            if sys in row1[1].value:
                file1.write(str(row1[39].value))
                file1.write('\n')
    

"""
with open(f_name, 'w') as file1:
    for row1 in ws.rows:
        if row1[1].value == 'ERP':
            file1.write(row1[39].value)
            file1.write('\n')
"""