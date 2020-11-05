# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""
#--------------------------------------------------------------
#  DQVT v x.xx
#   
#  NO | 테이블 명 | 엔터티 명 | 컬럼 명 | 속성 명 | Null 여부 |
#  (0) (1)        (2)        (3)       (4)      (5)
# 
#  데이터 타입 | 길이 | DEFAULT | 컬럼 순서 | PK | FK
#  (6)         (7)    (8)       (9)        (10)  (11)
#--------------------------------------------------------------

import openpyxl
from openpyxl.styles import PatternFill, Color
import pandas as pd

wb = openpyxl.load_workbook("D:/Doc/IT_audit/감리수행 프로젝트 현황/2020/통합온라인쇼핑몰 구축 프로젝트/산출물/.bak/테이블정의서/hangaram_테이블정의서_search_20200708.xlsx")

ws = wb.get_sheet_by_name("검색DB")

print("max row = ", ws.max_row)
print("max col = ", ws.max_column)

whiteFill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
grayFill = PatternFill(start_color='EEEEEE', end_color='EEEEEE', fill_type='solid')
pinkFill = PatternFill(start_color='FFDDFF', end_color='FFDDFF', fill_type='solid')
greenFill = PatternFill(start_color='AAEEAA', end_color='AAEEAA', fill_type='solid')

wb_r = openpyxl.Workbook()
ws_r0 = wb_r.create_sheet(index=0, title='동일컬럼속성명불일치')
ws_r1 = wb_r.create_sheet(index=1, title='동일속성명컬럼ID불일치')
ws_r2 = wb_r.create_sheet(index=2, title='PK없는테이블')
ws_r3 = wb_r.create_sheet(index=3, title='동일컬럼데이터타입불일치')
ws_r4 = wb_r.create_sheet(index=4, title='동일컬럼데이터길이불일치')
ws_r5 = wb_r.create_sheet(index=5, title='여부컬럼데이터길이검증')
    
def func_동일컬럼속성명불일치검색(ws_r0) :
    result = {}
    result = set()
    for row1 in ws.rows:
        if row1[3].value is not None:
            prev_colname = ""
            a = ""
            s = str(row1[3].value)
            for row in ws.rows:
                a=str(row[3].value)
                if a == s :
        
                    if len(prev_colname) > 0 and prev_colname != row[4].value :
                        result.add(s)
                        break
                    else :
                        prev_colname = str(row[4].value)
    
    ws_r0.cell(1,1, 'No').fill = greenFill
    ws_r0.cell(1,2, '테이블 명').fill = greenFill
    ws_r0.cell(1,3, '엔터티 명').fill = greenFill
    ws_r0.cell(1,4, '컬럼 명').fill = greenFill
    ws_r0.cell(1,5, '속성 명').fill = greenFill
    ws_r0.cell(1,6, 'Null 여부').fill = greenFill
    ws_r0.cell(1,7, '데이터 타입').fill = greenFill
    ws_r0.cell(1,8, '길이').fill = greenFill
    ws_r0.cell(1,9, 'DEFAULT').fill = greenFill
    ws_r0.cell(1,10, '컬럼순서').fill = greenFill
    ws_r0.cell(1,11, 'PK').fill = greenFill
    ws_r0.cell(1,12, 'FK').fill = greenFill
    
    j=2
    m=1
    
    for i in range(len(result)):
        b = str(result.pop())
        m = m * -1
        for row in ws.rows:
            if row[3].value is not None:
                a=str(row[3].value)
                if b == a :
                    for k in range(12) :
                        ws_r0.cell(j, k+1, row[k].value)
                        if k == 4: 
                            ws_r0.cell(j, k+1).fill = pinkFill
                        else :    
                            if m < 0 :
                                ws_r0.cell(j, k+1).fill = whiteFill
                            else :
                                ws_r0.cell(j, k+1).fill = grayFill
                    j = j+1

def func_동일속성명컬럼ID불일치검색(ws_r1) :
    result = {}
    result = set()
    for row1 in ws.rows:
        if row1[4].value is not None:
            prev_colname = ""
            a = ""
            s = str(row1[4].value)
            for row in ws.rows:
                if row[4].value is not None:
                    a=str(row[4].value)
                    if a == s :
            
                        if len(prev_colname) > 0 and prev_colname != row[3].value :
                            result.add(s)
                            break
                        else :
                            prev_colname = str(row[3].value)
    
    ws_r1.cell(1,1, 'No').fill = greenFill
    ws_r1.cell(1,2, '테이블 명').fill = greenFill
    ws_r1.cell(1,3, '엔터티 명').fill = greenFill
    ws_r1.cell(1,4, '컬럼 명').fill = greenFill
    ws_r1.cell(1,5, '속성 명').fill = greenFill
    ws_r1.cell(1,6, 'Null 여부').fill = greenFill
    ws_r1.cell(1,7, '데이터 타입').fill = greenFill
    ws_r1.cell(1,8, '길이').fill = greenFill
    ws_r1.cell(1,9, 'DEFAULT').fill = greenFill
    ws_r1.cell(1,10, '컬럼순서').fill = greenFill
    ws_r1.cell(1,11, 'PK').fill = greenFill
    ws_r1.cell(1,12, 'FK').fill = greenFill
    
    
    j=2
    m=1
    
    for i in range(len(result)):
        b = str(result.pop())
        m = m * -1
        for row in ws.rows:
            if row[4].value is not None:
                a=str(row[4].value)
                if b == a :
                    for k in range(12) :
                        ws_r1.cell(j, k+1, row[k].value)
                        if k == 3: 
                            ws_r1.cell(j, k+1).fill = pinkFill
                        else :    
                            if m < 0 :
                                ws_r1.cell(j, k+1).fill = whiteFill
                            else :
                                ws_r1.cell(j, k+1).fill = grayFill
                    j = j+1         


str_PK = 'PRI'
def func_PK없는테이블(ws_r2) :
    result = {}
    result = set()
    pk_yn = False
    prev_tbl = ""
    for row1 in ws.rows:
        if row1[1].value is not None:
            s = str(row1[1].value)  # 테이블명
            a = str(row1[10].value)  # PK
            if len(prev_tbl) > 0 and prev_tbl != s :
                if pk_yn == False :
                    result.add(prev_tbl)
                else :
                    pk_yn = False
                
            if str_PK in a : # PK표기가 있으면
                pk_yn = True
            
            prev_tbl = s
    
    ws_r2.cell(1,1, 'No').fill = greenFill
    ws_r2.cell(1,2, '테이블 명').fill = greenFill
    ws_r2.cell(1,3, '엔터티 명').fill = greenFill
    ws_r2.cell(1,4, '컬럼 명').fill = greenFill
    ws_r2.cell(1,5, '속성 명').fill = greenFill
    ws_r2.cell(1,6, 'Null 여부').fill = greenFill
    ws_r2.cell(1,7, '데이터 타입').fill = greenFill
    ws_r2.cell(1,8, '길이').fill = greenFill
    ws_r2.cell(1,9, 'DEFAULT').fill = greenFill
    ws_r2.cell(1,10, '컬럼순서').fill = greenFill
    ws_r2.cell(1,11, 'PK').fill = greenFill
    ws_r2.cell(1,12, 'FK').fill = greenFill
    
    
    j=2
    m=1
    
    for i in range(len(result)):
        b = str(result.pop())
        m = m * -1
        for row in ws.rows:
            if row[1].value is not None:
                a=str(row[1].value)
                if b == a :
                    for k in range(12) :
                        ws_r2.cell(j, k+1, row[k].value)
                        if k == 10: 
                            ws_r2.cell(j, k+1).fill = pinkFill
                        else :    
                            if m < 0 :
                                ws_r2.cell(j, k+1).fill = whiteFill
                            else :
                                ws_r2.cell(j, k+1).fill = grayFill
                    j = j+1

def func_동일컬럼데이터타입불일치검색(ws_r3) :
    result = {}
    result = set()
    for row1 in ws.rows:
        if row1[3].value is not None:
            prev_colname = ""
            a = ""
            s = str(row1[3].value)
            for row in ws.rows:
                if row[3].value is not None:
                    a=str(row[3].value)
                    if a == s :
                        if len(prev_colname) > 0 and prev_colname != row[6].value :               
                            result.add(s)
                            break
                        else :
                            prev_colname = str(row[6].value)
    
    ws_r3.cell(1,1, 'No').fill = greenFill
    ws_r3.cell(1,2, '테이블 명').fill = greenFill
    ws_r3.cell(1,3, '엔터티 명').fill = greenFill
    ws_r3.cell(1,4, '컬럼 명').fill = greenFill
    ws_r3.cell(1,5, '속성 명').fill = greenFill
    ws_r3.cell(1,6, 'Null 여부').fill = greenFill
    ws_r3.cell(1,7, '데이터 타입').fill = greenFill
    ws_r3.cell(1,8, '길이').fill = greenFill
    ws_r3.cell(1,9, 'DEFAULT').fill = greenFill
    ws_r3.cell(1,10, '컬럼순서').fill = greenFill
    ws_r3.cell(1,11, 'PK').fill = greenFill
    ws_r3.cell(1,12, 'FK').fill = greenFill
    
    j=2
    m=1
    
    for i in range(len(result)):
        b = str(result.pop())
        m = m * -1
        for row in ws.rows:
            if row[3].value is not None:
                a=str(row[3].value)
                if b == a :
                    for k in range(12) :
                        ws_r3.cell(j, k+1, row[k].value)
                        if k == 6 : 
                            ws_r3.cell(j, k+1).fill = pinkFill
                        else :    
                            if m < 0 :
                                ws_r3.cell(j, k+1).fill = whiteFill
                            else :
                                ws_r3.cell(j, k+1).fill = grayFill
                    j = j+1

def func_동일컬럼데이터길이불일치검색(ws_r4) :
    result = {}
    result = set()
    for row1 in ws.rows:
        if row1[3].value is not None:
            prev_colname = ""
            a = ""
            s = str(row1[3].value)
            for row in ws.rows:
                if row[3].value is not None:
                    a=str(row[3].value)
                    if a == s :
                        if len(prev_colname) > 0 and prev_colname != str(row[7].value) :               
                            result.add(s)
                            break
                        else :
                            prev_colname = str(row[7].value)
    
    ws_r4.cell(1,1, 'No').fill = greenFill
    ws_r4.cell(1,2, '테이블 명').fill = greenFill
    ws_r4.cell(1,3, '엔터티 명').fill = greenFill
    ws_r4.cell(1,4, '컬럼 명').fill = greenFill
    ws_r4.cell(1,5, '속성 명').fill = greenFill
    ws_r4.cell(1,6, 'Null 여부').fill = greenFill
    ws_r4.cell(1,7, '데이터 타입').fill = greenFill
    ws_r4.cell(1,8, '길이').fill = greenFill
    ws_r4.cell(1,9, 'DEFAULT').fill = greenFill
    ws_r4.cell(1,10, '컬럼순서').fill = greenFill
    ws_r4.cell(1,11, 'PK').fill = greenFill
    ws_r4.cell(1,12, 'FK').fill = greenFill
    
    j=2
    m=1
    
    for i in range(len(result)):
        b = str(result.pop())
        m = m * -1
        for row in ws.rows:
            if row[3].value is not None:
                a=str(row[3].value)
                if b == a :
                    for k in range(12) :
                        ws_r4.cell(j, k+1, row[k].value)
                        if k == 7: 
                            ws_r4.cell(j, k+1).fill = pinkFill
                        else :    
                            if m < 0 :
                                ws_r4.cell(j, k+1).fill = whiteFill
                            else :
                                ws_r4.cell(j, k+1).fill = grayFill
                    j = j+1

def func_여부컬럼데이터길이검증(ws_r5) :
    
    ws_r5.cell(1,1, 'No').fill = greenFill
    ws_r5.cell(1,2, '테이블 명').fill = greenFill
    ws_r5.cell(1,3, '엔터티 명').fill = greenFill
    ws_r5.cell(1,4, '컬럼 명').fill = greenFill
    ws_r5.cell(1,5, '속성 명').fill = greenFill
    ws_r5.cell(1,6, 'Null 여부').fill = greenFill
    ws_r5.cell(1,7, '데이터 타입').fill = greenFill
    ws_r5.cell(1,8, '길이').fill = greenFill
    ws_r5.cell(1,9, 'DEFAULT').fill = greenFill
    ws_r5.cell(1,10, '컬럼순서').fill = greenFill
    ws_r5.cell(1,11, 'PK').fill = greenFill
    ws_r5.cell(1,12, 'FK').fill = greenFill
    
    j=2
    m=1
    for row1 in ws.rows:
        if (row1[3].value is not None) and (row1[4].value is not None) :
            s = str(row1[3].value)
            a = str(row1[4].value)
            if 'YN' in s or '여부' in a or '성별' in a :
                if row1[7].value is not None:
                    if int(str(row1[7].value)) != 1 :
                        m = m * -1
                        for k in range(12) :
                            ws_r5.cell(j, k+1, row1[k].value)
                            if k == 7: 
                                ws_r5.cell(j, k+1).fill = pinkFill
                            else :    
                                if m < 0 :
                                    ws_r5.cell(j, k+1).fill = whiteFill
                                else :
                                    ws_r5.cell(j, k+1).fill = grayFill
                        j = j+1

#--------------------------------------------------------------
# 실행 (DQVT_result.xlsx)
#
#--------------------------------------------------------------

func_동일컬럼속성명불일치검색(ws_r0)
func_동일속성명컬럼ID불일치검색(ws_r1)
func_PK없는테이블(ws_r2)
func_동일컬럼데이터타입불일치검색(ws_r3)
func_동일컬럼데이터길이불일치검색(ws_r4)    
func_여부컬럼데이터길이검증(ws_r5)

wb_r.save("C:/Users/Pulmuone/Desktop/DQVT_result_search_20200708.xlsx")
