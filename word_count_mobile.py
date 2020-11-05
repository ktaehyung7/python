# -*- coding: utf-8 -*-
"""
Created on Thu Nov  5 16:27:00 2020

@author: Pulmuone
"""

import openpyxl
from openpyxl.styles import PatternFill, Color
import pandas as pd
from konlpy.tag import Twitter
from collections import Counter

wb = openpyxl.load_workbook("D:/Doc/12. IT 시스템 만족도 조사/IT만족도조사_exam/4_survey_all_2020(최종)v1.1.xlsx")

ws = wb.get_sheet_by_name("Sheet2")
ws2 = wb.get_sheet_by_name("모바일의견WC")


def get_tags(text, ntags=50):
    spliter = Twitter()
    # konlpy의 Twitter객체
    nouns = spliter.nouns(text)
    # nouns 함수를 통해서 text에서 명사만 분리/추출
    count = Counter(nouns)
    # Counter객체를 생성하고 참조변수 nouns할당
    return_list = []  # 명사 빈도수 저장할 변수
    for n, c in count.most_common(ntags):
        temp = {'tag': n, 'count': c}
        return_list.append(temp)
    # most_common 메소드는 정수를 입력받아 객체 안의 명사중 빈도수
    # 큰 명사부터 순서대로 입력받은 정수 갯수만큼 저장되어있는 객체 반환
    # 명사와 사용된 갯수를 return_list에 저장합니다.
    return return_list
 
 
def main():
    sys_list = ['ebills', 'ERP', 'FMS', 'FNC', 'KWP', 'SFA', '톡톡']
    i=2
    for sys in sys_list:
        f_name = "D:/Doc/12. IT 시스템 만족도 조사/IT만족도조사_exam/R_data/모바일의견/원본/모바일의견("+sys+").txt"
        with open(f_name, 'w') as file1:
            for row1 in ws.rows:
                if sys in row1[1].value:
                    file1.write(str(row1[39].value))
                    file1.write('\n')
        file1.close()
        
        #text_file_name = "D:/Doc/12. IT 시스템 만족도 조사/IT만족도조사_exam/R_data/기타의견/원본/기타의견(ebills).txt"
        # 분석할 파일
        noun_count = 9999
        # 최대 많은 빈도수 부터 20개 명사 추출
        output_file_name = "D:/Doc/12. IT 시스템 만족도 조사/IT만족도조사_exam/R_data/모바일의견/결과/result("+sys+")_m.txt"
        # count.txt 에 저장
        open_text_file = open(f_name, 'r',-1,"ANSI")
        # 분석할 파일을 open 
        text = open_text_file.read() #파일을 읽습니다.
        tags = get_tags(text, noun_count) # get_tags 함수 실행
        open_text_file.close()   #파일 close
        open_output_file = open(output_file_name, 'w',-1,"ANSI")
        # 결과로 쓰일 count.txt 열기
        for tag in tags:
            noun = tag['tag']
            count = tag['count']
            open_output_file.write('{} {}\n'.format(noun, count))
            # 결과 저장
        open_output_file.close() 
        
        #f_name2 = "D:/Doc/12. IT 시스템 만족도 조사/IT만족도조사_exam/R_data/모바일의견/결과/result("+sys+")_m.txt"
    
        with open(output_file_name, 'r') as file2:
            lines = file2.readlines()
    
            for line in lines:
                line_str=line.split(' ')
                if 'KWP' in sys:
                    sys="KWP/O365"
                ws2.cell(i, 1, sys)
                ws2.cell(i, 2, line_str[0])
                ws2.cell(i, 3, line_str[1])
                i=i+1
        
        file2.close()

    wb.save("D:/Doc/12. IT 시스템 만족도 조사/IT만족도조사_exam/4_survey_all_2020(최종)v1.1.xlsx")
        
if __name__ == '__main__':
    main()