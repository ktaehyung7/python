# -*- coding: utf-8 -*-
"""
Created on Thu Nov  5 15:06:39 2020

@author: Pulmuone
"""

import openpyxl
from openpyxl.styles import PatternFill, Color
import pandas as pd



wb = openpyxl.load_workbook("D:/Doc/12. IT 시스템 만족도 조사/IT만족도조사_exam/4_survey_all_2020(최종)v1.1.xlsx")

ws = wb.get_sheet_by_name("Sheet2")

print("max row = ", ws.max_row)
print("max col = ", ws.max_column)

sys_list = ['ebills', 'ERP', 'FDD PHI', 'FMS', 'FNC', 'KWP', 'LDS PHI', 'MRS', 'OMS(식품)', 'PIPS', 'PLM', 'SFA', '톡톡']

for sys in sys_list:
    f_name = "D:/Doc/12. IT 시스템 만족도 조사/IT만족도조사_exam/R_data/기타의견/원본/기타의견("+sys+").txt"
    with open(f_name, 'w') as file1:
        for row1 in ws.rows:
            if sys in row1[1].value:
                file1.write(str(row1[38].value))
                file1.write('\n')